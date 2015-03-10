from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter

wb = Workbook()
ws = wb.active

ws.title = 'New Shiny Worksheet'
for col_idx in range(1, 40):
	col = get_column_letter(col_idx)
	for row in range(1, 600):
		ws.cell('%s%s'%(col, row)).value = '%s%s' % (col, row)

ws = wb.create_sheet()
ws.title = 'Second worksheet'
ws['F5'] = 'F5'
dest_filename = 'test2.xlsx'
wb.save(filename = dest_filename)