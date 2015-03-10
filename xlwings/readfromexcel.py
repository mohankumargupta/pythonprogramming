from openpyxl import load_workbook

wb = load_workbook(filename = 'test1.xlsx')
#print(dir(wb))
worksheet  = wb.get_sheet_by_name('Sheet1')
print(worksheet['A1'].value)
print(worksheet['B1'].value)
print('-----------------')
for row in worksheet.iter_rows():
	for cell in row:
	    print(cell.value) 